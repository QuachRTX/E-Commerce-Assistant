__author__ = "Alexsander Stahnke"
__version__ = "1.6"
__status__ = "Beta"
__email__ = "alexsander.stahnke@grendene.com.br"
__github__ = "https://github.com/QuachRTX"

from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
import customtkinter as ctk


def converter_string_para_numero(valor):
    try:
        valor = valor.strip()
        if valor.isdigit():
            return int(valor)
        valor_sem_pontos = valor.replace('.', '').replace(',', '.')
        return float(valor_sem_pontos)
    except ValueError:
        return valor

def converter_celulas_para_numeros(caminho_planilha):
    wb = load_workbook(filename=caminho_planilha)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    valor_numerico = converter_string_para_numero(cell.value)
                    if valor_numerico != cell.value:
                        cell.value = valor_numerico
    wb.save(caminho_planilha)

def mapear_colunas(sheet, colunas_esperadas):
    mapeamento = {}
    cabeçalho = [cell.value.lower() if isinstance(cell.value, str) else cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    for col_esperada, variacoes in colunas_esperadas.items():
        variacoes = [v.lower() for v in variacoes]
        for i, col_nome in enumerate(cabeçalho):
            if col_nome in variacoes:
                mapeamento[col_esperada] = i
                break
    return mapeamento

def carregar_dados_planilha(caminho, colunas_esperadas):
    wb = load_workbook(filename=caminho)
    sheet = wb.active
    mapeamento_colunas = mapear_colunas(sheet, colunas_esperadas)
    dados = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = {}
        for col_interna, idx in mapeamento_colunas.items():
            valor = row[idx]
            item[col_interna] = valor
        dados.append(item)
    return dados

def validar_ajustes_de_preco(dados_base, dados_referencia):
    try:
        resultados = []
        produtos_base = {item_base['produto'] for item_base in dados_base}

        for item_base in dados_base:
            correspondencias = [
                item_referencia for item_referencia in dados_referencia
                if item_referencia['produto'] == item_base['produto'] and item_referencia['cor'] == item_base['cor']
            ]
            for correspondencia in correspondencias:
                if item_base['preço de'] != correspondencia['preço de'] or item_base['preço por'] != correspondencia['preço por']:
                    resultado = {
                        "Produto": item_base['produto'],
                        "Cor": item_base['cor'],
                        "Status": "Preços divergentes",
                        "Encontrado (Planilha Base)": f"De: {item_base['preço de']}, Por: {item_base['preço por']}",
                        "Esperado (Planilha Atualizada)": f"De: {correspondencia['preço de']}, Por: {correspondencia['preço por']}"
                    }
                    resultados.append(resultado)

        # Verificar produtos em dados_referencia que não estão em dados_base
        for item_referencia in dados_referencia:
            if item_referencia['produto'] not in produtos_base:
                resultado = {
                    "Produto": item_referencia['produto'],
                    "Cor": item_referencia.get('cor', 'N/A'),  # 'N/A' se 'Cor' não estiver disponível
                    "Status": "Não encontrado",
                    "Encontrado (Planilha Base)": " ",
                    "Esperado (Planilha Atualizada)": f"de: {item_referencia.get('preço de', ' ')}, por: {item_referencia.get('preço por', ' ')}"
                }
                resultados.append(resultado)

        return resultados
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro durante a validação: {e}")
        return None  # Retorna None para indicar um erro

def exportar_divergencias(resultados, nome_arquivo):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Preços Divergentes'
    ws.append(['Produto', 'Cor', 'Status', 'Encontrado (Planilha Base)', 'Esperado (Planilha Atualizada)'])
    for resultado in resultados:
        ws.append([
            resultado['Produto'],
            resultado['Cor'],
            resultado['Status'],
            resultado['Encontrado (Planilha Base)'],
            resultado['Esperado (Planilha Atualizada)']
        ])
    wb.save(nome_arquivo)

class PriceValidatorApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Validador de Preços")
        self.geometry("600x400")

        # Selecionar Planilha do Gestor (Base)
        ctk.CTkLabel(self, text="Selecione a Planilha do Gestor (Base):").pack(pady=10)
        self.btn_planilha1 = ctk.CTkButton(self, text="Abrir Planilha 1", command=lambda: self.abrir_planilha('1'))
        self.btn_planilha1.pack(pady=5)
        self.label_planilha1 = ctk.CTkLabel(self, text="")

        # Selecionar Planilha Atualizada (Referência)
        ctk.CTkLabel(self, text="Selecione a Planilha Atualizada (Referência):").pack(pady=10)
        self.btn_planilha2 = ctk.CTkButton(self, text="Abrir Planilha 2", command=lambda: self.abrir_planilha('2'))
        self.btn_planilha2.pack(pady=5)

        # Botão para baixar o modelo de planilha
        self.btn_baixar_modelo = ctk.CTkButton(self, text="Baixar modelo de planilha", command=self.criar_modelo_planilha)
        self.btn_baixar_modelo.pack(pady=5)
        self.btn_baixar_modelo.configure(fg_color='transparent', text_color='red')  # Correta transparência e texto em vermelho

        # Botão Validar Preços
        self.btn_validar = ctk.CTkButton(self, text="Validar Preços", command=self.validar_precos, state="disabled", width=150, height=35)
        self.btn_validar.pack(pady=50)

        # Botão de Ajuda
        self.btn_ajuda = ctk.CTkButton(self, text="Ajuda!?", command=self.mostrar_ajuda)
        self.btn_ajuda.pack(pady=10, anchor='se', side='right')

        self.planilha1_path = ''
        self.planilha2_path = ''

    def mostrar_ajuda(self):
        mensagem_ajuda = ("Caso encontre problemas, verifique se os nomes das colunas na 'Planilha2' estão corretos.\n"
                          "Recomendo fortemente a remover edições de cores e fontes de texto!\n"
                          "Utilize os seguintes nomes de colunas caso encontre problemas:\n"
                          "- Produto\n"
                          "- Cor\n"
                          "- Preço De\n"
                          "- Preço Por\n\n"
                          "Esses nomes ajudam a garantir que o aplicativo funcione corretamente! :)")
        messagebox.showinfo("Ajuda", mensagem_ajuda)

    def abrir_planilha(self, numero):
        path = filedialog.askopenfilename(title=f"Abrir Planilha {numero}", filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            if numero == '1':
                self.planilha1_path = path
                self.btn_planilha1.configure(text="Planilha Selecionada")
            elif numero == '2':
                self.planilha2_path = path
                self.btn_planilha2.configure(text="Planilha Selecionada")
            self.checar_selecao()

    def criar_modelo_planilha(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Modelo"
        colunas = ["Produto", "Cor", "Preço De", "Preço Por"]
        ws.append(colunas)
        nome_arquivo = f'Modelo Referência.xlsx'
        nome_arquivo = filedialog.asksaveasfilename(
            title="Salvar Modelo de Planilha",
            filetypes=[("Excel files", "*.xlsx")],
            defaultextension=".xlsx", initialfile=nome_arquivo
        )
        if nome_arquivo:
            wb.save(nome_arquivo)
            messagebox.showinfo("Sucesso", "Modelo de planilha criado e salvo com sucesso em: " + nome_arquivo)
        else:
            messagebox.showinfo("Aviso", "Operação cancelada. Modelo de planilha não foi salvo.")

    def checar_selecao(self):
        if self.planilha1_path and self.planilha2_path:
            self.btn_validar.configure(state="normal")

    def validar_precos(self):
        converter_celulas_para_numeros(self.planilha1_path)
        converter_celulas_para_numeros(self.planilha2_path)
    
        colunas_esperadas = {
            'produto': ['Produto', 'Código', 'Codigo', 'Sku', 'Código do produto', 'Cód'],
            'cor': ['Cor', 'Cód cor', 'Cód_cor', 'Cod_Cor', 'sku_cor', 'Código da Cor'],
            'preço de': ['Preço De', 'preço de', 'Preço de', 'NOVO PREÇO', 'Preço DE'],
            'preço por': ['Preço Por', 'preço por', 'Preço por', 'NOVO PREÇO POR', 'Preço POR atual', 'Preço POR']
        }
        dados_planilha1 = carregar_dados_planilha(self.planilha1_path, colunas_esperadas)
        dados_planilha2 = carregar_dados_planilha(self.planilha2_path, colunas_esperadas)
        
        resultados_validacao = validar_ajustes_de_preco(dados_planilha1, dados_planilha2)
        
        if resultados_validacao is None:
            # Não fazer nada adicional, uma mensagem de erro já foi exibida pela função validar_ajustes_de_preco
            return
        elif not resultados_validacao:
            messagebox.showinfo("Validação", "Não foram encontradas divergências.")
        else:
            nome_arquivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if nome_arquivo:
                exportar_divergencias(resultados_validacao, nome_arquivo)
                messagebox.showinfo("Concluído", "Arquivo de divergências salvo com sucesso.")


# Dentro de checkpriceadjusts.py, ajustado conforme a nova definição de PriceValidatorApp
def main():
    app = PriceValidatorApp()
    app.mainloop()